from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import pytesseract
from PIL import Image
import re
from simulated_data import simulated_dataset
from sklearn.tree import DecisionTreeClassifier
import numpy as np
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io

app = Flask(__name__)
# Update CORS to allow Vercel frontend (adjust for production)
CORS(app, resources={r"/*": {"origins": ["http://localhost:3000", "https://*.vercel.app"]}})  # Allow local and Vercel
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure upload folder exists (though Vercel uses /tmp for serverless)
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Train ML model
X = [row[:4] for row in simulated_dataset]
y = [row[4] for row in simulated_dataset]
model = DecisionTreeClassifier()
model.fit(X, y)

def parse_blood_parameters(text):
    # Updated regex patterns to be more flexible with whitespace, case, and formatting
    patterns = {
        'hemoglobin': r'(?:hemoglobin|hb|hgb)[ :]*([\d.]+)\s*(?:g/dL|g\/dL)?',
        'rbc': r'(?:RBC count|red blood cell)[ :]*([\d.]+)\s*(?:million/uL|million\/µL)?',
        'wbc': r'(?:WBC count|white blood cell)[ :]*([\d]+)\s*(?:/uL|\/µL)?',
        'platelet': r'(?:platelet count|plt)[ :]*([\d]+)\s*(?:/uL|\/µL)?'
    }
    parameters = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            parameters[key] = float(match.group(1)) if '.' in match.group(1) else int(match.group(1))
    return parameters

def get_reference_ranges(age, gender):
    age = int(age)
    ranges = {
        'hemoglobin': {'low': 13.0, 'high': 17.0} if gender == 'male' else {'low': 12.0, 'high': 15.0},
        'rbc': {'low': 4.5, 'high': 5.9} if gender == 'male' else {'low': 4.1, 'high': 5.1},
        'wbc': {'low': 4000, 'high': 11000},
        'platelet': {'low': 150000, 'high': 450000}
    }
    if age < 18:
        ranges['hemoglobin']['low'] -= 1.0
        ranges['hemoglobin']['high'] -= 1.0
    return ranges

def compare_to_ranges(parameters, ranges):
    status = {}
    for param, value in parameters.items():
        if param in ranges:
            low = ranges[param]['low']
            high = ranges[param]['high']
            if value < low:
                status[param] = 'Low'
            elif value > high:
                status[param] = 'High'
            else:
                status[param] = 'Normal'
    return status

# Updated dietary recommendations for each condition
def get_dietary_recommendations(condition):
    if condition == "Normal":
        return [
            "Maintain a balanced diet with iron-rich foods like spinach and red meat.",
            "Stay hydrated with 8–10 glasses of water daily.",
            "Include vitamin C-rich foods (e.g., oranges) to boost iron absorption.",
            "Eat whole grains like quinoa for sustained energy.",
            "Incorporate healthy fats from nuts and avocados for overall wellness."
        ]
    elif condition == "Anemia":
        return [
            "Increase iron intake with foods like liver, spinach, and lentils.",
            "Consume vitamin B12-rich foods such as eggs and fish.",
            "Add folate-rich foods like broccoli and beans to your diet.",
            "Avoid tea or coffee with meals, as they inhibit iron absorption.",
            "Consider iron supplements after consulting a doctor."
        ]
    elif condition == "Infection":
        return [
            "Boost immunity with vitamin C-rich foods like citrus fruits and bell peppers.",
            "Include zinc-rich foods such as pumpkin seeds and oysters.",
            "Eat probiotic-rich foods like yogurt to support gut health.",
            "Stay hydrated with electrolyte drinks or coconut water.",
            "Avoid processed foods to reduce inflammation."
        ]
    elif condition == "Danger":
        return [
            "Urgently increase iron and protein intake with foods like beef and eggs.",
            "Consume leafy greens like kale for folate and iron.",
            "Add vitamin B12 sources like dairy or fortified cereals.",
            "Avoid alcohol and caffeine to prevent further nutrient depletion.",
            "Seek immediate medical attention for potential blood transfusions."
        ]
    elif condition == "Critical":
        return [
            "Limit sodium intake to reduce blood pressure (e.g., avoid processed foods).",
            "Eat potassium-rich foods like bananas to support heart health.",
            "Include omega-3 fatty acids from fish like salmon to reduce inflammation.",
            "Avoid sugary foods to prevent blood sugar spikes.",
            "Consult a cardiologist immediately for heart attack risk assessment."
        ]
    return ["Consult a doctor for personalized advice."]

@app.route('/')
def home():
    return "VitalEdge Backend"

@app.route('/analyze', methods=['POST'])
def analyze():
    age = request.form.get('age')
    gender = request.form.get('gender')
    symptoms = request.form.get('symptoms', '')
    file = request.files.get('file')

    if not age or not gender:
        return jsonify({'error': 'Missing age or gender'}), 400

    if not file:
        return jsonify({'error': 'No file uploaded'}), 400

    filename = file.filename
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)

    try:
        img = Image.open(file_path)
        text = pytesseract.image_to_string(img)
    except Exception as e:
        return jsonify({'error': f'OCR failed: {str(e)}'}), 500

    parameters = parse_blood_parameters(text)
    ranges = get_reference_ranges(age, gender)
    status = compare_to_ranges(parameters, ranges)

    required_params = ['hemoglobin', 'rbc', 'wbc', 'platelet']
    if not all(param in parameters for param in required_params):
        return jsonify({
            "user_info": {"age": age, "gender": gender, "symptoms": symptoms},
            "analysis": {
                "file_path": file_path,
                "extracted_text": text,
                "blood_parameters": parameters,
                "reference_ranges": ranges,
                "parameter_status": status,
                "predicted_condition": "Insufficient data for prediction",
                "dietary_recommendations": ["Please ensure all parameters (hemoglobin, RBC, WBC, platelet) are present in the report"]
            },
            "status": "error",
            "timestamp": "2025-03-18"
        }), 200

    input_data = [parameters.get(param, 0) for param in required_params]

    # New condition logic based on status counts
    status_counts = {'Low': 0, 'Normal': 0, 'High': 0}
    for param_status in status.values():
        status_counts[param_status] += 1

    if status_counts['Normal'] == 4:
        condition = "Normal"
    elif status_counts['Low'] == 4:
        condition = "Danger"
    elif status_counts['High'] >= 2:
        condition = "Critical"
    else:
        # Mixed cases: Use Decision Tree for nuanced prediction
        condition = model.predict([input_data])[0]

    recommendations = get_dietary_recommendations(condition)

    response = {
        "user_info": {"age": age, "gender": gender, "symptoms": symptoms},
        "analysis": {
            "file_path": file_path,
            "extracted_text": text,
            "blood_parameters": parameters,
            "reference_ranges": ranges,
            "parameter_status": status,
            "predicted_condition": condition,
            "dietary_recommendations": recommendations
        },
        "status": "success",
        "timestamp": "2025-03-18"
    }
    return jsonify(response), 200

@app.route('/download-excel', methods=['POST'])
def download_excel():
    data = request.get_json()
    analysis = data['analysis']

    wb = Workbook()
    ws = wb.active
    ws.title = "VitalEdge Analysis"

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    headers = ['Parameter', 'Value', 'Status', 'Range Low', 'Range High']
    ws.append(headers)

    for row, (key, value) in enumerate(analysis['blood_parameters'].items(), start=2):
        status = analysis['parameter_status'].get(key, 'N/A')
        range_low = analysis['reference_ranges'].get(key, {}).get('low', 'N/A')
        range_high = analysis['reference_ranges'].get(key, {}).get('high', 'N/A')
        ws.append([key, value, status, range_low, range_high])
        
        cell = ws.cell(row=row, column=3)
        if status == 'Low':
            cell.fill = red_fill
        elif status == 'High':
            cell.fill = yellow_fill
        elif status == 'Normal':
            cell.fill = green_fill

    ws.append(['Predicted Condition', analysis['predicted_condition']])
    ws.append(['Recommendations', '; '.join(analysis['dietary_recommendations'])])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='VitalEdge_Analysis.xlsx', as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True, port=5000)